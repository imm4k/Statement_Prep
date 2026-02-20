from __future__ import annotations

import sys
from pathlib import Path
from typing import List

from openpyxl import Workbook

_THIS_DIR = Path(__file__).resolve().parent
_ROOT_DIR = _THIS_DIR.parent
_PART2_DIR = _ROOT_DIR / "part2_ppt"

for p in (str(_PART2_DIR), str(_ROOT_DIR), str(_THIS_DIR)):
    if p not in sys.path:
        sys.path.insert(0, p)

import config
from excel_inputs import load_general_config
from ppt_objects import UpdateContext, apply_ownership_amount
from ppt_monthly_stmt_values import (
    list_investor_owner_property_triplets,
    build_monthly_perf_totals,
    build_monthly_cash_totals,
)

def export_monthly_stmt_excel() -> Path:
    setup_xlsx = Path(config.SETUP_EXCEL_PATH)
    general = load_general_config(
        xlsx_path=setup_xlsx,
        sheet_name=config.GENERAL_CONFIG_SHEET,
        label_output_location=config.GENERAL_CONFIG_LABEL_OUTPUT_LOCATION,
        label_statement_thru=config.GENERAL_CONFIG_LABEL_STATEMENT_THRU_DATE,
    )

    yyyy_mm = general.statement_thru_date.strftime("%Y_%m")
    out_path = general.output_location / f"{yyyy_mm}_monthly_stmt_data.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    ws_perf = wb.create_sheet("monthly_perf_table")
    ws_cash = wb.create_sheet("monthly_cash_table")

    perf_cols = [
        "Investor",
        "Owner",
        "Property",
        "Rent",
        "Dividend",
        "Total Revenue",
        "HOA & Mgt. Fee",
        "Repairs & Other Exp.",
        "Mortgage Interest",
        "Total Expenses",
        "Monthly",
        "Cumulative",
    ]

    cash_cols = [
        "Investor",
        "Owner",
        "Property",
        "Owner Contribution",
        "Mortgage Loan",
        "Rent & Dividend",
        "Total Inflow",
        "HOA & Mgt. Fee",
        "Repairs & Other Exp.",
        "Mortgage Interest",
        "Mortgage Principal",
        "Apartment & Improve.",
        "Owner Distribution",
        "Total Outflow",
        "Monthly",
        "Cumulative",
    ]

    ws_perf.append(perf_cols)
    ws_cash.append(cash_cols)

    triplets = list_investor_owner_property_triplets()
    total = len(triplets)

    print(f"Found {total} Investor/Owner/Property combinations to export.")
    if total == 0:
        print("No data found in gl_agg. Export will contain headers only.")

    for idx, (investor, owner, prop) in enumerate(triplets, start=1):

        if idx == 1 or idx % 25 == 0 or idx == total:
            print(f"Currently on {idx} of {total}")
        ctx = UpdateContext(
            investor=investor,
            owner=owner,
            ownership_pct=100.0,
            ownership_factor=1.0,
            statement_thru_date_dt=general.statement_thru_date,
            statement_thru_date_str=general.statement_thru_date.strftime("%m/%d/%Y"),
            t1_str=general.statement_thru_date.strftime("%b %Y"),
        )

        perf = build_monthly_perf_totals(ctx, property_name=prop)
        cash = build_monthly_cash_totals(ctx, property_name=prop)

        perf_row: List[object] = [investor, owner, prop]
        for k in perf_cols[3:]:
            v = float(perf.get(k, 0.0))
            v_out = apply_ownership_amount(ctx, v, f"export.monthly_perf_table.{k}")
            perf_row.append(v_out)
        ws_perf.append(perf_row)

        cash_row: List[object] = [investor, owner, prop]
        for k in cash_cols[3:]:
            v = float(cash.get(k, 0.0))
            v_out = apply_ownership_amount(ctx, v, f"export.monthly_cash_table.{k}")
            cash_row.append(v_out)
        ws_cash.append(cash_row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    print(f"Saved monthly statement export: {out_path}")
    return out_path


def main() -> None:
    export_monthly_stmt_excel()


if __name__ == "__main__":
    main()
