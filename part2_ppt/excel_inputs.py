from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import config
from openpyxl import load_workbook

@dataclass(frozen=True)
class GeneralConfig:
    statement_thru_date: datetime
    output_location: Path


def load_general_config(xlsx_path: Path, sheet_name: str, label_output_location: str, label_statement_thru: str) -> GeneralConfig:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{sheet_name}' in setup workbook.")
    ws = wb[sheet_name]

    output_location = _find_label_value(ws, label_output_location)
    statement_thru = _find_label_value(ws, label_statement_thru)

    if output_location is None or str(output_location).strip() == "":
        raise ValueError("Could not locate Output Location value in General Config.")
    if statement_thru is None:
        raise ValueError("Could not locate Statement Thru Date value in General Config.")

    statement_thru_dt = _coerce_to_datetime(statement_thru)
    return GeneralConfig(statement_thru_date=statement_thru_dt, output_location=Path(str(output_location)))

@dataclass(frozen=True)
class RunConfigRow:
    investor: str
    owner: str
    standard_slides_template: str

def load_run_config_rows(xlsx_path: Path, sheet_name: str) -> List[RunConfigRow]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{sheet_name}' in setup workbook.")
    ws = wb[sheet_name]

    header_row_idx, investor_col_idx = _find_header_cell(ws, "Investor")
    if header_row_idx is None or investor_col_idx is None:
        raise ValueError("Could not find an 'Investor' header in Run Config.")

    _, owner_col_idx = _find_header_cell(ws, "Owner")
    if owner_col_idx is None:
        raise ValueError("Could not find an 'Owner' header in Run Config.")

    _, base_template_col_idx = _find_header_cell(ws, "Base Template")
    if base_template_col_idx is None:
        _, base_template_col_idx = _find_header_cell(ws, "Template")

    rows: List[RunConfigRow] = []
    seen_pairs = set()

    row_idx = header_row_idx + 1
    while True:
        investor_val = ws.cell(row=row_idx, column=investor_col_idx).value
        if investor_val is None or str(investor_val).strip() == "":
            break

        owner_val = ws.cell(row=row_idx, column=owner_col_idx).value
        if owner_val is None or str(owner_val).strip() == "":
            raise ValueError(f"Run Config row {row_idx} has an Investor but missing Owner.")

        investor = str(investor_val).strip()
        owner = str(owner_val).strip()

        pair_key = (investor.lower(), owner.lower())
        if pair_key in seen_pairs:
            print(f"Run Config duplicate Investor Owner pair detected and skipped. Row {row_idx}. Investor: {investor}. Owner: {owner}")
            row_idx += 1
            continue
        seen_pairs.add(pair_key)

        tmpl_val = ""
        if base_template_col_idx is not None:
            v = ws.cell(row=row_idx, column=base_template_col_idx).value
            tmpl_val = "" if v is None else str(v).strip()

        rows.append(
            RunConfigRow(
                investor=investor,
                owner=owner,
                standard_slides_template=tmpl_val,
            )
        )

        row_idx += 1

    if not rows:
        raise ValueError("No run rows found under the 'Investor' column in Run Config.")

    return rows

def load_investor_table_ownership_map(xlsx_path: Path) -> Dict[Tuple[str, str], List[float]]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    try:
        sheet_name = getattr(config, "INVESTOR_TABLE_SHEET", "Investor Table")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet '{sheet_name}' in setup workbook.")
        ws = wb[sheet_name]

        header_row_idx, investor_col_idx = _find_header_cell(ws, "Investor")
        if header_row_idx is None or investor_col_idx is None:
            raise ValueError("Could not find an 'Investor' header in Investor Table.")

        _, owner_col_idx = _find_header_cell(ws, "Owner")
        if owner_col_idx is None:
            raise ValueError("Could not find an 'Owner' header in Investor Table.")

        _, property_col_idx = _find_header_cell(ws, "Property")
        if property_col_idx is None:
            raise ValueError("Could not find a 'Property' header in Investor Table.")

        _, ownership_col_idx = _find_header_cell(ws, "% Ownership")
        if ownership_col_idx is None:
            raise ValueError("Could not find a '% Ownership' header in Investor Table.")

        def _parse_pct_100_format(value: object, row_idx_local: int) -> float:
            if value is None:
                raise ValueError(f"Investor Table row {row_idx_local} missing % Ownership.")

            if isinstance(value, str):
                s = value.strip()
                if s == "":
                    raise ValueError(f"Investor Table row {row_idx_local} missing % Ownership.")
                s = s.replace("%", "").strip()
                try:
                    v = float(s)
                except Exception:
                    raise ValueError(f"Investor Table row {row_idx_local} has invalid % Ownership: {value}")
                return v

            try:
                v_num = float(value)
            except Exception:
                raise ValueError(f"Investor Table row {row_idx_local} has invalid % Ownership: {value}")

            if 0.0 <= v_num <= 1.0:
                return v_num * 100.0

            return v_num

        ownership_by_inv_owner: Dict[Tuple[str, str], List[float]] = {}
        sums_by_owner_property: Dict[Tuple[str, str], float] = {}

        row_idx = header_row_idx + 1
        while True:
            investor_val = ws.cell(row=row_idx, column=investor_col_idx).value
            if investor_val is None or str(investor_val).strip() == "":
                break

            owner_val = ws.cell(row=row_idx, column=owner_col_idx).value
            property_val = ws.cell(row=row_idx, column=property_col_idx).value

            if owner_val is None or str(owner_val).strip() == "":
                raise ValueError(f"Investor Table row {row_idx} has an Investor but missing Owner.")
            if property_val is None or str(property_val).strip() == "":
                raise ValueError(f"Investor Table row {row_idx} has an Investor but missing Property.")

            investor = str(investor_val).strip()
            owner = str(owner_val).strip()
            prop = str(property_val).strip()

            pct_val = ws.cell(row=row_idx, column=ownership_col_idx).value
            pct = _parse_pct_100_format(pct_val, row_idx)

            if pct <= 0.0 or pct > 100.0:
                raise ValueError(
                    f"Investor Table row {row_idx} has % Ownership out of range: {pct}. Valid range is >0 and <=100."
                )

            inv_owner_key = (investor.lower(), owner.lower())
            if inv_owner_key not in ownership_by_inv_owner:
                ownership_by_inv_owner[inv_owner_key] = []
            ownership_by_inv_owner[inv_owner_key].append(float(pct))

            owner_prop_key = (owner.lower(), prop.lower())
            sums_by_owner_property[owner_prop_key] = sums_by_owner_property.get(owner_prop_key, 0.0) + float(pct)

            row_idx += 1

        if not ownership_by_inv_owner:
            raise ValueError("No Investor Table rows found under the 'Investor' column.")

        tolerance = 0.01
        bad = []
        for (owner_k, prop_k), total in sums_by_owner_property.items():
            if abs(float(total) - 100.0) > tolerance:
                bad.append((owner_k, prop_k, float(total)))

        if bad:
            sample = bad[:10]
            lines = "; ".join([f"Owner={o} Property={p} Sum={t}" for o, p, t in sample])
            raise ValueError(
                f"Investor Table ownership sums must equal 100% for each Owner Property. "
                f"Found {len(bad)} failures. Sample: {lines}"
            )

        return ownership_by_inv_owner
    finally:
        wb.close()

def _find_label_value(ws, label_text: str) -> Optional[object]:
    for row in ws.iter_rows(values_only=False):
        cell = row[0]
        if cell.value is None:
            continue
        if str(cell.value).strip() == label_text:
            return ws.cell(row=cell.row, column=cell.column + 1).value
    return None


def _find_header_cell(ws, header_text: str) -> Tuple[Optional[int], Optional[int]]:
    target = header_text.strip().lower()
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, min(ws.max_column, 100) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().lower() == target:
                return r, c
    return None, None


def _coerce_to_datetime(value: object) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        raise ValueError(f"Unrecognized date string format for Statement Thru Date: {value}")
    raise ValueError(f"Unsupported Statement Thru Date value type: {type(value)}")
